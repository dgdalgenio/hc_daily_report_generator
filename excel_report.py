"""
excel_report.py
----------------
Builds the styled Part 2 "City Assignment Summary" Excel workbook and
converts it into HTML for embedding in the email report.
"""

import io

import openpyxl
import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from xlsx2html import xlsx2html


def build_city_pivot_workbook(report_df_dict: dict, totals_df: pd.DataFrame,
                               location_to_dss_dict: dict, cities_lst: list, report_date):
    """Builds the styled workbook and returns (workbook, raw_xlsx_bytes, excel_html_content)."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "City Assignment Summary"
    ws.views.sheetView[0].showGridLines = True

    font_family, thin_side = "Segoe UI", Side(border_style="thin", color="D9D9D9")
    timestamp_font = Font(name=font_family, size=10, italic=True, color="555555")
    label_font = Font(name=font_family, size=11, bold=True, color="333333")
    value_font = Font(name=font_family, size=11, color="1B365D")
    header_font = Font(name=font_family, size=11, bold=True, color="1B365D")
    data_font = Font(name=font_family, size=11, color="333333")
    header_fill = PatternFill(start_color="EEECE1", end_color="EEECE1", fill_type="solid")
    all_borders = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    ws.cell(row=1, column=1, value=f"Report Generated: {report_date}").font = timestamp_font
    current_row = 3

    ws.cell(row=current_row, column=1, value="Overall Summary Totals:").font = Font(
        name=font_family, size=12, bold=True, color="1B365D"
    )
    current_row += 1

    if not totals_df.empty:
        for col_idx, col_name in enumerate(["City/Location"] + list(totals_df.columns), start=1):
            c = ws.cell(row=current_row, column=col_idx, value=col_name)
            c.font, c.fill, c.border = header_font, header_fill, all_borders
            c.alignment = Alignment(horizontal="left" if col_idx == 1 else "center", vertical="center")
        current_row += 1

        for index_val, row_values in zip(totals_df.index, totals_df.values):
            combined_row = [index_val] + list(row_values)
            r_font = label_font if str(index_val).strip() == 'Total' else data_font
            for col_idx, val in enumerate(combined_row, start=1):
                data_cell = ws.cell(row=current_row, column=col_idx, value=val)
                data_cell.font, data_cell.border = r_font, all_borders
                if col_idx == 1:
                    data_cell.alignment = Alignment(horizontal="left", vertical="center")
                else:
                    data_cell.value = int(val) if str(val).isdigit() else val
                    data_cell.number_format = "#,##0"
                    data_cell.alignment = Alignment(horizontal="center", vertical="center")
            current_row += 1

    current_row += 2

    for city in cities_lst:
        if city not in report_df_dict:
            continue
        ws.cell(row=current_row, column=1, value="Assigned City:").font = label_font
        c_val = ws.cell(row=current_row, column=2, value=city)
        c_val.font = Font(name=font_family, size=11, bold=True, color="1B365D")
        current_row += 1

        ws.cell(row=current_row, column=1, value="Assigned DSS:").font = label_font
        d_val = ws.cell(row=current_row, column=2, value=location_to_dss_dict.get(city, {}).get('name', 'N/A'))
        d_val.font = value_font
        current_row += 1

        df_city = report_df_dict[city]
        for col_idx, col_name in enumerate(df_city.columns, start=1):
            hc = ws.cell(row=current_row, column=col_idx, value=col_name)
            hc.font, hc.fill, hc.border = header_font, header_fill, all_borders
            hc.alignment = Alignment(horizontal="left" if col_idx == 1 else "center", vertical="center")
        current_row += 1

        for row_data in df_city.values:
            is_total_row = (row_data[0] == 'Total')
            row_font = label_font if is_total_row else data_font
            for col_idx, val in enumerate(row_data, start=1):
                dc = ws.cell(row=current_row, column=col_idx, value=val)
                dc.font, dc.border = row_font, all_borders
                if col_idx == 1:
                    dc.alignment = Alignment(horizontal="left", vertical="center")
                else:
                    dc.value = int(val) if str(val).isdigit() else val
                    dc.number_format = "#,##0"
                    dc.alignment = Alignment(horizontal="center", vertical="center")
            current_row += 1
        current_row += 1

    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value and cell.row > 1:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max(max_len + 5, 18)

    xlsx_buffer = io.BytesIO()
    wb.save(xlsx_buffer)
    xlsx_buffer.seek(0)
    raw_xlsx_bytes = xlsx_buffer.getvalue()

    excel_html_buffer = io.StringIO()
    xlsx2html(xlsx_buffer, excel_html_buffer)
    excel_html_content = excel_html_buffer.getvalue()

    return wb, raw_xlsx_bytes, excel_html_content